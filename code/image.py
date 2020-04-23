import requests
from openpyxl import workbook
from openpyxl import load_workbook

def download_img(img_url,index):
    print (img_url)
    r = requests.get(img_url)
    print(r.status_code) # 返回状态码
    if r.status_code == 200:
        open('./img/' +str(index) +'.jpg', 'wb').write(r.content) # 将内容写入图片
        print("done")
    del r

def get_URL():
    wb = load_workbook('url.xlsx')
    sheet = wb.get_sheet_by_name('Sheet')
    maxR = sheet.max_row
    maxC = sheet.max_column
    # print(maxC)
    # print(maxR)
    url = []
    for r in range(1,maxR+1):
        # print(r)
        for c in range(2, maxC+1):
            # print(c)
            if sheet.cell(row=r,column=c).value is None:
                continue
            else:
                url.append(sheet.cell(row=r,column=c).value)
            # print(sheet.cell(row=r,column=c).value)


    wb.close()
    return url


if __name__ == '__main__':
    # 下载要的图片
    # img_url = "https://wx2.sinaimg.cn/orj480/86a17422ly1gd7lo2e542j21hc0u078p.jpg"
    # download_img(img_url)
    # print(get_URL())
    url = get_URL()
    index = 1
    for u in url:
        download_img(u,index)
        index = index+1


